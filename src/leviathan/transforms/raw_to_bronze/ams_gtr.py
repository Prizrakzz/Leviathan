"""Bronze transform for the USDA AMS Grain Transportation Report (GTR) freight family.

Converts the raw payloads this family lands in S3 -- AgTransport SODA JSON and the
one ``GTRTable1.xlsx`` spreadsheet -- into typed, source-faithful pandas DataFrames.

No S3 or AWS dependencies: pure data transformation.  The fetcher, the Batch task
and the tests all call these functions directly.

This module is the SINGLE AUTHORITY on what the family contains.  :data:`GTR_DATASETS`
names every dataset, its channel, its endpoint, its cadence and -- most importantly --
its UNIT, together with the verbatim phrase the source itself uses to declare that
unit.  The fetcher reads this table to know what to request; the silver transform
reads it to know what a row means.  Nothing in the family is spelled twice.

Channels and why both exist
---------------------------
The lane asked for the SODA JSON API to be preferred over the spreadsheets, and it
is -- for six of the seven datasets.  The exception is measured, not assumed:

    GTRTable1.xlsx sheet ``Data`` is the ONLY publication of the WEEKLY
    dollars-per-metric-ton Gulf->Japan and PNW->Japan ocean rate.

The recon recorded that a "SODA twin exists" for that series and pointed at
``8uye-ieij`` (Grain Transportation Cost Indicators).  That is wrong, and the way it
is wrong is a units trap worth stating in full because ``8uye-ieij``'s own column
metadata invites it.  Its columns ``gulf_vessel`` / ``pacific_vessel`` are described
by the publisher as "Ocean freight rate from the U.S. Gulf to Japan ($/metric ton)",
but the DATASET description says the file holds "weekly cost indices ... the base of
each index (set to 100) is each mode's average cost in the year 2017".  Measured
live on 2026-08-20 for the same report week:

    8uye-ieij   gulf_vessel     = 184.9733028222731   (index, 2017 = 100)
    GTRTable1   Ocean Gulf      =  72.75              (US dollars per metric ton)

The index value reproduces GTRTable1's own index cell to full float precision, so
``8uye-ieij`` is the INDEX twin of Table 1, not the rate twin.  A fetcher that
believed the column description would have published index points under a
dollars-per-metric-ton name -- a units lie no downstream consumer could detect.
``8uye-ieij`` is therefore REFUSED by this family (see :data:`REFUSED_DATASETS`),
and the weekly ocean rate is taken from the spreadsheet.

The unit of that spreadsheet column is not asserted from faith either.  GTRTable1
declares no unit anywhere in the sheet, so the weekly series is cross-asserted
against ``ehs5-yac3`` (Vessel Rates), whose column metadata says verbatim "The rates
from the U.S. Gulf to Japan in dollar per metric ton".  Over the twelve most recent
overlapping months the weekly mean and the published monthly value agree to within
2.8% (Gulf) and 2.0% (PNW) -- same series, same unit.  ``tests/unit/
test_transforms_ams_gtr.py`` pins that agreement on fixtures.

Host access -- MEASURED, and the recon is overturned here too
--------------------------------------------------------------
The recon states that ``ams.usda.gov`` "403s non-browser UAs" and that any fetcher
"must set a browser User-Agent".  Re-probed on 2026-08-20 against
``GTRTable1.xlsx``:

    (no UA header, python-urllib default)          -> 200, 297,236 bytes
    "leviathan-etl/1.0"                            -> 200, 297,236 bytes
    "Leviathan-GTR/1.0"                            -> 200, 297,236 bytes
    requests default ("python-requests/x.y")       -> 200, 297,236 bytes
    "Leviathan-Ingest/1.0 (+research data pipeline; contact <addr>)" -> 403

So the host is NOT browser-gated.  What draws the 403 is a particular UA SHAPE --
the long parenthetical-with-contact form -- not the absence of a browser string.
A short honest product token passes.  This family therefore sends a short honest
token and NEVER a fake browser UA.  The AgTransport SODA endpoints do not gate on
UA at all (every dataset below answered 200 to python-urllib's default).

PIT: the knowledge date, and where derivation stops
---------------------------------------------------
GTR publishes weekly on THURSDAY.  Neither channel states which report week
published a given row, so silver derives a release date per the D-LD
derived-release-date idiom -- derive, LABEL the derivation, never hide it.  The
derivation rule is per-dataset and lives in :attr:`GtrDataset.knowledge_basis`:

``derived_gtr_thursday``
    Weekly datasets.  ``knowledge_date`` = the first Thursday strictly after
    ``period_date``.  Measured period-date weekdays: GTRTable1 is Wednesday for all
    1,253 rows; the four barge datasets are Tuesday (1,039 dates) or Wednesday (142).
    Residual, stated rather than hidden: when a Thursday is a federal holiday AMS may
    publish on the Friday, which would make this derivation one day EARLY.  The basis
    label is carried on every row so a consumer can add a buffer.

``derived_gtr_thursday_month_end``
    ``ocean_monthly``.  ``period_date`` is the FIRST of the month (measured: all 367
    dates are day 1), so the derivation runs off the month END -- deriving off the
    period_date would claim a July figure was knowable on 2 July.

``derived_ams_ukraine_annual_edition``
    ``ukraine_ocean_quarterly``.  Its SODA metadata declares "Publishing Schedule:
    Yearly" and links the AMS Ukraine Grain Transportation annual report, so a
    Thursday derivation is meaningless: quarter-end 2025-12-31 would claim knowledge
    on 2026-01-01 for a figure that appears in an annual edition months later.  The
    release date is instead the first annual EDITION strictly after the quarter end,
    taken at the edition month's LAST day (the conservative direction, since only the
    edition month is known).  All seven editions in
    :data:`AMS_UKRAINE_ANNUAL_EDITIONS` were verified 200 live on 2026-08-20.  Their
    ``Last-Modified`` headers are NOT usable as release dates -- six of the seven read
    2025-09-15, a site-migration re-stamp.  When no edition follows the quarter, the
    derivation has no answer and the row falls back to the OBSERVED snapshot date.

``observed_snapshot``
    The fallback everywhere: ``knowledge_date`` = ``as_of_date``, the date the bytes
    were actually fetched.  Always true, never early.

Every derived date is checked against the snapshot: a row present in a snapshot taken
at ``as_of`` was published on or before ``as_of``, so ``knowledge_date <= as_of_date``
must hold.  The silver transform raises when it does not -- that inequality failing is
the tell that a derivation has stopped describing reality.

Missing measures stay NULL (INV-4)
----------------------------------
A rate that the source does not quote is absent, not zero.  Measured absences:
``deqi-uken`` has 774 null rates of 8,267 rows (9.4%), ``7spn-fbua`` 2,076 of 30,706
(6.8%), and GTRTable1's ocean columns carry 27 nulls of 1,253 written as the literal
string ``n/a``.  All become NULL and none becomes 0.0.

A zero is not a quote either
----------------------------
The absence machinery above only fires on NON-numeric text, so a source-side zero used
as a missing marker walks straight through it -- the same failure INV-4 exists to stop,
one step upstream, and worse than a NULL because a 0.0 is silently averaged, min'd and
plotted.  This family contains exactly one, and it is unambiguous.

Measured live 2026-08-20 across every rate column in the family (``$where=<col> = 0``
on all six SODA endpoints): ONE record, ``7spn-fbua`` (2025-12-02,
"La Crosse - Minneapolis"), publishes ``price_per_ton: "0"``.  It is the FIRST week of
that reach's seasonal ice closure on the upper Mississippi, and the publisher spells
the other fifteen weeks of the SAME closure -- 2025-12-09 through 2026-03-17 -- as
genuine absences with no ``price_per_ton`` key at all.  The prior winter is absences
throughout with no zero anywhere (2024-11-26 onward).  On the week itself all 26
reaches are served and 25 carry real rates, so it is that reach's shutdown and not a
file-wide gap.  A $0.00/ton barge rate would in any case mean the barge moved free.

RULE, therefore: a literal zero in a rate column is refused as a quote and becomes NULL
with a WARN naming the count (:func:`_null_zero_sentinels`).  The narrow, principled
exemption is :data:`ZERO_IS_A_QUOTE_COLUMNS` -- today just ``gulf_pnw_spread``, a
DIFFERENCE between two routes rather than a rate, for which zero is an ordinary market
state.  The raw object keeps the zero regardless; only its promotion to a number is
refused.

Unrecognised cell text is surfaced, not swallowed
-------------------------------------------------
GTRTable1 row 236 (period 2007-01-03) holds the string ``One week Lag`` in the Ocean
Gulf value cell.  Tokens on :data:`_MISSING_TOKENS` become NULL silently; ANY other
non-numeric string becomes NULL with a WARN naming the token, so a future annotation
is caught rather than lost (the INV-1 schema-drift idiom).

Licence
-------
Public domain.  The AMS datasets page states verbatim: "These data series are
aggregated from non-confidential and non-copyrighted sources."  USDA policy requests
attribution as "U.S. Department of Agriculture".  ``ocean_monthly`` additionally
carries "SOURCE: O'Neil Commodity Consulting" -- a private vendor whose attribution
AMS prints and which this family carries through to every row in
``source_attribution`` rather than dropping at the silver boundary.
"""
from __future__ import annotations

import datetime as _dt
import json
from dataclasses import dataclass

import pandas as pd

from leviathan.common.logging import get_logger

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Units -- the closed vocabulary
# ---------------------------------------------------------------------------
# Three units appear in this family and they are NOT interchangeable.  A
# percent-of-tariff number (~600-800) and a dollars-per-ton number (~14-17) describe
# the same barge move; averaging or plotting them together is meaningless.  Silver
# carries the unit on EVERY row for exactly that reason.
UNIT_USD_PER_METRIC_TON = "USD_per_metric_ton"
UNIT_USD_PER_TON = "USD_per_ton"
UNIT_PCT_OF_TARIFF = "percent_of_tariff"

# The closed set, used to pin that no dataset invents a fourth unit.
KNOWN_UNITS = frozenset({UNIT_USD_PER_METRIC_TON, UNIT_USD_PER_TON, UNIT_PCT_OF_TARIFF})

# Attribution strings carried through to silver.
ATTRIBUTION_USDA = "U.S. Department of Agriculture"
ATTRIBUTION_USDA_ONEIL = (
    "U.S. Department of Agriculture; SOURCE: O'Neil Commodity Consulting"
)

# ---------------------------------------------------------------------------
# Refusals -- written down rather than left as an absence
# ---------------------------------------------------------------------------
REFUSED_DATASETS: dict[str, str] = {
    "8uye-ieij": (
        "Grain Transportation Cost Indicators. REFUSED as a rate source: its "
        "gulf_vessel / pacific_vessel column descriptions claim '($/metric ton)' but "
        "the dataset description declares cost INDICES with 2017 = 100, and the "
        "measured value for report week 2026-08-19 is 184.9733028222731 against "
        "GTRTable1's 72.75 USD/mt for the same week -- it reproduces Table 1's INDEX "
        "cell to full float precision. Publishing it under a USD/mt name would be an "
        "undetectable units lie. If the index itself is ever wanted it must land as "
        "its own dataset with unit='index_2017_100', which is a schema decision, not "
        "a rename."
    ),
    "nb3e-9fyt": (
        "Ukraine Grain Transportation Cost Indicators. Created 2026-01-09 and still "
        "column-less; /resource/nb3e-9fyt.json answers 403 (Socrata's response for a "
        "non-tabular asset). Nothing to fetch yet -- watch it."
    ),
    "hwhq-eta9,bwaz-8sgs,965a-yzgy,si5y-wpqw": (
        "Bulk Grain Ocean Dashboard / Bulk Vessel Fleet Size and Rates / Barge "
        "Dashboard x2. All answer 403 on /resource/{id}.json: they are dashboard "
        "assets, not tabular datasets, and have no SODA row endpoint."
    ),
}

# Cell strings that mean "not quoted".  Anything else non-numeric gets a WARN.
_MISSING_TOKENS = frozenset({"", "n/a", "na", "n.a.", "-", "--", "nan", "null"})

# THE ZERO RULE (measured 2026-08-20; see the module docstring, "A zero is not a quote").
# A literal 0 in a freight RATE column is refused as a quote and becomes NULL with a WARN.
# These are the only columns in the family where an exact zero is a real value, and each is
# listed because of what it MEASURES, never because a zero was observed in it:
#
#   gulf_pnw_spread -- a DIFFERENCE (US Gulf-Japan less PNW-Japan), not a rate. Zero means the
#       two routes priced level, which is an ordinary market state. It has not happened in the
#       367 months 1996-01..2026-07 (measured live: `$where=gulf_pnw_spread = 0` -> 0 rows), but
#       it could tomorrow, and nulling it would delete a real observation.
#
# Every other value column here is a price or a percent-of-tariff, where zero would mean the
# barge or the vessel moved for free. Add to this set only with a measurement that shows the
# publisher means zero.
ZERO_IS_A_QUOTE_COLUMNS = frozenset({"gulf_pnw_spread"})

# ---------------------------------------------------------------------------
# The AMS Ukraine Grain Transportation annual report editions
# ---------------------------------------------------------------------------
# Used ONLY to derive a release date for ``ukraine_ocean_quarterly``.  Each entry is
# (year, month) of the edition as AMS names the file; the derived release date is the
# LAST day of that month, because only the month is known and later is the safe
# direction.  Every URL below returned 200 on 2026-08-20 (HEAD, plain requests UA).
AMS_UKRAINE_ANNUAL_EDITIONS: tuple[tuple[int, int, str], ...] = (
    (2020, 3, "UkraineMarch2020.pdf"),
    (2021, 4, "UkraineApril2021.pdf"),
    (2022, 8, "UkraineAugust2022.pdf"),
    (2023, 6, "UkraineJune2023.pdf"),
    (2024, 6, "UkraineJune2024.pdf"),
    (2025, 6, "Ukraine%20June%202025.pdf"),
    (2026, 7, "Ukraine_Grain_Transportation_2026_July.pdf"),
)

# Knowledge-date derivation bases.
BASIS_GTR_THURSDAY = "derived_gtr_thursday"
BASIS_GTR_THURSDAY_MONTH_END = "derived_gtr_thursday_month_end"
BASIS_UKRAINE_ANNUAL = "derived_ams_ukraine_annual_edition"
BASIS_OBSERVED = "observed_snapshot"


# ---------------------------------------------------------------------------
# The dataset table
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class GtrDataset:
    """One dataset of the GTR freight family.

    Attributes:
        slug:            Estate-side dataset name; also the S3 ``dataset=`` segment
                         and the silver partition value.
        channel:         ``"soda"`` or ``"xlsx"``.
        endpoint:        SODA four-by-four id, or the absolute spreadsheet URL.
        filename:        Raw payload filename under the dataset prefix.
        period_col:      Source column holding the period date.
        period_grain:    ``weekly`` | ``monthly`` | ``quarterly``.
        key_col:         Source column naming the route / river reach, if any.
        value_cols:      Source value columns -> silver ``series`` names.
        unit:            The unit EVERY value column carries.
        unit_declaration: Substring the source's own column metadata must contain for
                         the unit to be considered asserted.  Checked by
                         :func:`assert_soda_unit_declaration` against the live
                         ``/api/views/{id}.json`` payload captured beside the data.
        knowledge_basis: Which derivation rule silver applies.
        attribution:     Value for the silver ``source_attribution`` column.
        forward_month_offset: 0 for spot; 1 / 3 for the forward barge curves.
        extra_key_cols:  Additional source columns that are part of the natural key.
        notes:           Free text carried into the module's own documentation.
    """

    slug: str
    channel: str
    endpoint: str
    filename: str
    period_col: str
    period_grain: str
    value_cols: dict[str, str]
    unit: str
    unit_declaration: str
    knowledge_basis: str
    attribution: str = ATTRIBUTION_USDA
    key_col: str | None = None
    forward_month_offset: int = 0
    extra_key_cols: tuple[str, ...] = ()
    notes: str = ""


GTR_DATASETS: dict[str, GtrDataset] = {
    # -- (1) the weekly ocean rate: the lane's highest-value item -------------
    "ocean_weekly": GtrDataset(
        slug="ocean_weekly",
        channel="xlsx",
        endpoint="https://www.ams.usda.gov/sites/default/files/media/GTRTable1.xlsx",
        filename="GTRTable1.xlsx",
        period_col="period_date",
        period_grain="weekly",
        value_cols={
            "ocean_gulf_japan": "gulf_to_japan",
            "ocean_pnw_japan": "pnw_to_japan",
        },
        unit=UNIT_USD_PER_METRIC_TON,
        # GTRTable1 declares no unit in the sheet; the unit is cross-asserted against
        # ehs5-yac3, which declares it verbatim.  See the module docstring.
        unit_declaration="dollar per metric ton",
        knowledge_basis=BASIS_GTR_THURSDAY,
        notes=(
            "Sheet 'Data', header on row 7, data from row 8. 1,253 weekly rows "
            "2002-08-21 -> 2026-08-19, every period_date a Wednesday. Bronze keeps "
            "Table 1's diesel / rail / river columns too (source-faithful); silver "
            "emits ONLY the two ocean series, which is the lane's item (1). The "
            "other three are a deliberate silver-side hold, not an oversight: "
            "'river' duplicates deqi-uken's Illinois reach in different units, and "
            "diesel and rail are a different family (fuel and rail) that should land "
            "with their own contract rather than ride the freight table."
        ),
    ),
    # -- the monthly ocean rate: the SODA form, and the unit authority --------
    "ocean_monthly": GtrDataset(
        slug="ocean_monthly",
        channel="soda",
        endpoint="ehs5-yac3",
        filename="full.json",
        period_col="date",
        period_grain="monthly",
        value_cols={
            "gulf_to_japan": "gulf_to_japan",
            "pnw_to_japan": "pnw_to_japan",
            "gulf_pnw_spread": "gulf_pnw_spread",
        },
        unit=UNIT_USD_PER_METRIC_TON,
        unit_declaration="dollar per metric ton",
        knowledge_basis=BASIS_GTR_THURSDAY_MONTH_END,
        attribution=ATTRIBUTION_USDA_ONEIL,
        notes=(
            "367 monthly rows 1996-01-01 -> 2026-07-01, period_date always day 1. "
            "Attributed by AMS to O'Neil Commodity Consulting; the attribution "
            "travels on every row."
        ),
    ),
    # -- (2) the barge complex: four endpoints, two units --------------------
    "barge_pct_tariff": GtrDataset(
        slug="barge_pct_tariff",
        channel="soda",
        endpoint="deqi-uken",
        filename="full.json",
        period_col="date",
        period_grain="weekly",
        key_col="location",
        value_cols={"rate": "barge_rate"},
        unit=UNIT_PCT_OF_TARIFF,
        unit_declaration="percent-of-tariff",
        knowledge_basis=BASIS_GTR_THURSDAY,
        notes=(
            "8,267 rows 2004-01-07 -> 2026-08-18 over 7 origins (Twin Cities, "
            "Mid-Mississippi, Lower Illinois, St. Louis, Cincinnati, Lower Ohio, "
            "Cairo-Memphis). 774 null rates kept NULL."
        ),
    ),
    "barge_per_ton": GtrDataset(
        slug="barge_per_ton",
        channel="soda",
        endpoint="7spn-fbua",
        filename="full.json",
        period_col="date",
        period_grain="weekly",
        key_col="river_system_location",
        value_cols={"price_per_ton": "barge_price_per_ton"},
        unit=UNIT_USD_PER_TON,
        unit_declaration="price per ton",
        knowledge_basis=BASIS_GTR_THURSDAY,
        notes=(
            "30,706 rows 2004-01-07 -> 2026-08-18 over 26 named river reaches; "
            "2,076 null rates kept NULL. UNIT CONFLICT IN THE SOURCE, resolved on "
            "the column: the dataset TITLE says 'Per Ton Rates' and the column "
            "description says 'Price Per Ton calculated as the benchmark rate for "
            "the location', but the dataset DESCRIPTION opens 'Weekly barge rates "
            "(percent of tariff)' -- copied from deqi-uken. Magnitude decides it: "
            "values run 14-17 where a percent-of-tariff runs 600-800, so this is "
            "USD per ton. The column declaration is the authority; the dataset "
            "blurb is a copy-paste defect."
        ),
    ),
    "barge_fwd_1m": GtrDataset(
        slug="barge_fwd_1m",
        channel="soda",
        endpoint="svms-9yya",
        filename="full.json",
        period_col="date",
        period_grain="weekly",
        key_col="location",
        value_cols={"rate": "barge_rate"},
        unit=UNIT_PCT_OF_TARIFF,
        unit_declaration="percent-of-tariff",
        knowledge_basis=BASIS_GTR_THURSDAY,
        forward_month_offset=1,
        extra_key_cols=("rate_month",),
        notes=(
            "8,267 rows, the ONE-MONTH-FORWARD barge curve. 'date' is when the "
            "quote was COLLECTED; 'rate_month' is the calendar month the quote "
            "applies to and is part of the key."
        ),
    ),
    "barge_fwd_3m": GtrDataset(
        slug="barge_fwd_3m",
        channel="soda",
        endpoint="uuhv-5etw",
        filename="full.json",
        period_col="date",
        period_grain="weekly",
        key_col="location",
        value_cols={"rate": "barge_rate"},
        unit=UNIT_PCT_OF_TARIFF,
        unit_declaration="percent-of-tariff",
        knowledge_basis=BASIS_GTR_THURSDAY,
        forward_month_offset=3,
        extra_key_cols=("rate_month",),
        notes="8,267 rows, the THREE-MONTH-FORWARD barge curve. As barge_fwd_1m.",
    ),
    # -- (3) the direct Black Sea instrument ---------------------------------
    "ukraine_ocean_quarterly": GtrDataset(
        slug="ukraine_ocean_quarterly",
        channel="soda",
        endpoint="2n8s-739j",
        filename="full.json",
        period_col="quarter_ending_date",
        period_grain="quarterly",
        key_col="route",
        value_cols={"rate": "ukraine_ocean_rate"},
        unit=UNIT_USD_PER_METRIC_TON,
        unit_declaration="dollars per metric ton",
        knowledge_basis=BASIS_UKRAINE_ANNUAL,
        extra_key_cols=("commodity", "vessel_size"),
        notes=(
            "260 rows, 26 quarter-ends 2019-09-30 -> 2025-12-31, 10 routes out of "
            "Odesa / Mykolaiv / Izmail / Constanta, 3 commodities, 5 vessel sizes. "
            "The one direct Black Sea price in the family."
        ),
    ),
}

# Route labels for the ocean series, whose xlsx leg has no route column of its own.
# Spelled once, HERE, and imported by the silver transform: the weekly and monthly legs
# of the same route must join on route_or_reach, which two copies could silently break.
OCEAN_ROUTES: dict[str, str] = {
    "gulf_to_japan": "US Gulf-Japan",
    "pnw_to_japan": "PNW-Japan",
    "gulf_pnw_spread": "US Gulf-Japan less PNW-Japan",
}

# GTRTable1 sheet layout, asserted before a single value is read.
_XLSX_SHEET = "Data"
_XLSX_HEADER_ROW = 7
_XLSX_FIRST_DATA_ROW = 8
# 1-based column index -> (expected row-7 header text, bronze column name)
_XLSX_COLUMNS: dict[int, tuple[str, str]] = {
    1: ("Date", "period_date"),
    2: ("Price", "diesel_price_usd_per_gal"),
    3: ("Rail", "rail_usd_per_car"),
    4: ("River", "river_pct_of_tariff"),
    5: ("Gulf", "ocean_gulf_japan"),
    6: ("PNW", "ocean_pnw_japan"),
}
# Row 6 carries the group label sitting above the two ocean columns.
_XLSX_GROUP_CELL = (6, 5, "Ocean")  # (row, col, expected text)

SOURCE_NAME = "ams_gtr"


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def soda_resource_url(dataset: str) -> str:
    """Absolute SODA row endpoint for *dataset*.

    Raises:
        ValueError: If *dataset* is unknown or is not a SODA dataset.
    """
    spec = get_dataset(dataset)
    if spec.channel != "soda":
        raise ValueError(f"dataset {dataset!r} is channel={spec.channel!r}, not soda")
    return f"https://agtransport.usda.gov/resource/{spec.endpoint}.json"


def soda_metadata_url(dataset: str) -> str:
    """Absolute SODA metadata endpoint for *dataset* (the unit declaration lives here)."""
    spec = get_dataset(dataset)
    if spec.channel != "soda":
        raise ValueError(f"dataset {dataset!r} is channel={spec.channel!r}, not soda")
    return f"https://agtransport.usda.gov/api/views/{spec.endpoint}.json"


def get_dataset(dataset: str) -> GtrDataset:
    """Look up a dataset spec.  Fail-closed on an unknown slug."""
    try:
        return GTR_DATASETS[dataset]
    except KeyError:
        raise ValueError(
            f"unknown ams_gtr dataset {dataset!r}; known: {sorted(GTR_DATASETS)}"
        ) from None


def _to_float(value, *, unknown: set[str]) -> float | None:
    """Coerce one source cell to float, or ``None``.

    Args:
        value:   The raw cell.
        unknown: Accumulator for non-numeric text the caller should WARN about.

    Recognised missing tokens become ``None`` silently.  Any OTHER non-numeric string
    becomes ``None`` and is recorded in *unknown* so the caller can WARN once per
    token rather than once per row (INV-1 schema-drift idiom).  A missing measure is
    NEVER 0.0 (INV-4).
    """
    if value is None:
        return None
    if isinstance(value, bool):  # guard: bool is an int subclass
        unknown.add(repr(value))
        return None
    if isinstance(value, (int, float)):
        parsed = float(value)
        return None if parsed != parsed else parsed  # NaN -> None
    text = str(value).strip()
    if text.lower() in _MISSING_TOKENS:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        unknown.add(text)
        return None


def _to_date(value) -> _dt.date | None:
    """Coerce a source date cell (ISO string, Socrata timestamp, or datetime) to a date."""
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    # Socrata floating timestamps: "2026-08-19T00:00:00.000"
    return _dt.date.fromisoformat(text[:10])


def _null_zero_sentinels(values: pd.Series, dataset: str, column: str) -> pd.Series:
    """Refuse a literal zero as a freight quote: it becomes NULL, loudly.

    See :data:`ZERO_IS_A_QUOTE_COLUMNS` for the exemption list and the module docstring
    for the measurement this rule was written from.  Raw is immutable in S3, so the zero
    itself is never lost -- only its promotion to a number a mean or a min would swallow.

    Args:
        values:  Coerced float column.
        dataset: Dataset slug, for the log line.
        column:  SOURCE column name, matched against the exemption set.

    Returns:
        The column with exact zeros masked to NULL, or unchanged for an exempt column.
    """
    if column in ZERO_IS_A_QUOTE_COLUMNS:
        return values
    zeros = values == 0.0
    count = int(zeros.sum())
    if count:
        logger.warning(
            "ams_gtr %s: column %r holds %d literal ZERO value(s) -- each became NULL. "
            "A quoted freight rate is never 0.00; a source-side zero on a rate series is "
            "an absence the publisher spelled with a digit, and unlike a NULL it would be "
            "silently averaged, min'd and plotted (INV-4, one step upstream). If a zero "
            "here is ever a real quote, add the column to ZERO_IS_A_QUOTE_COLUMNS with the "
            "measurement that shows it.",
            dataset, column, count,
        )
    return values.mask(zeros)


def _warn_unknown_tokens(dataset: str, column: str, unknown: set[str]) -> None:
    if unknown:
        logger.warning(
            "ams_gtr %s: column %r holds %d unrecognised non-numeric token(s) %s -- "
            "each became NULL (never 0.0, INV-4). Raw is immutable in S3, so nothing "
            "is lost; promote a token to _MISSING_TOKENS only after deciding it means "
            "'not quoted'.",
            dataset, column, len(unknown), sorted(unknown)[:10],
        )


# ---------------------------------------------------------------------------
# Unit assertion against the source's own metadata
# ---------------------------------------------------------------------------

def assert_soda_unit_declaration(meta_bytes: bytes, dataset: str) -> str:
    """Assert that the source still declares the unit this family assumes.

    The fetcher captures ``/api/views/{id}.json`` beside every SODA payload precisely
    so that the unit is read from the publisher rather than remembered by us.  This
    checks that at least one of the dataset's value columns still carries
    :attr:`GtrDataset.unit_declaration` in its own description.

    Args:
        meta_bytes: Raw bytes of the ``/api/views/{id}.json`` response.
        dataset:    Dataset slug.

    Returns:
        The matching column description, for logging.

    Raises:
        ValueError: If the declaration is gone.  That is a units-drift alarm, not a
            nuisance: it means the publisher has restated what the numbers mean and
            the mapping in :data:`GTR_DATASETS` can no longer be trusted.
    """
    spec = get_dataset(dataset)
    meta = json.loads(meta_bytes)
    columns = {c.get("fieldName"): (c.get("description") or "") for c in meta.get("columns", [])}

    wanted = spec.unit_declaration.lower()
    for source_col in spec.value_cols:
        description = columns.get(source_col, "")
        if wanted in description.lower():
            logger.info(
                "ams_gtr %s: unit %s asserted from the source's own column metadata "
                "(%s: %r)", dataset, spec.unit, source_col, description[:120],
            )
            return description

    raise ValueError(
        f"ams_gtr {dataset}: the source no longer declares {spec.unit_declaration!r} "
        f"on any of {sorted(spec.value_cols)}. Observed descriptions: "
        f"{ {k: v[:80] for k, v in columns.items() if k in spec.value_cols} }. "
        "The publisher has restated what these numbers mean -- re-read the metadata "
        "and re-decide the unit before writing another row."
    )


# ---------------------------------------------------------------------------
# SODA JSON -> bronze
# ---------------------------------------------------------------------------

def transform_gtr_soda_json_to_bronze(
    raw_bytes: bytes,
    dataset: str,
    as_of_date: str,
    ingest_date: str,
) -> pd.DataFrame:
    """Parse one raw SODA JSON payload into a typed, source-faithful bronze frame.

    Args:
        raw_bytes:  Raw bytes of the SODA ``/resource/{id}.json`` response (a JSON array).
        dataset:    Dataset slug from :data:`GTR_DATASETS`.
        as_of_date: Snapshot date, ``YYYYMMDD``.
        ingest_date: ISO date (``YYYY-MM-DD``) this row was written.

    Returns:
        DataFrame carrying the source's own column names plus the bronze metadata
        columns.  Column ORDER is the spec's, not the payload's.

    Raises:
        ValueError: If the payload is not a non-empty JSON array, or if a column the
            dataset spec depends on is absent from every record.
    """
    spec = get_dataset(dataset)
    if spec.channel != "soda":
        raise ValueError(f"dataset {dataset!r} is channel={spec.channel!r}, not soda")

    records = json.loads(raw_bytes)
    if not isinstance(records, list):
        raise ValueError(
            f"ams_gtr {dataset}: SODA response is not a JSON array "
            f"({type(records).__name__})"
        )
    if not records:
        raise ValueError(
            f"ams_gtr {dataset}: SODA response is an empty array -- no data to transform."
        )

    df = pd.DataFrame(records)

    # Schema-drift alert (INV-1): raw is immutable in S3, so an unknown key is already
    # preserved there.  Surface it so a future AgTransport column addition is promoted
    # deliberately instead of vanishing at the typed projection.
    known = (
        {spec.period_col}
        | set(spec.value_cols)
        | set(spec.extra_key_cols)
        | ({spec.key_col} if spec.key_col else set())
        # Socrata's derived calendar helpers, present on most datasets.
        | {"week", "month", "year", "year_quarter", "quarter", "origin", "destination"}
    )
    unknown_fields = sorted(set(df.columns) - known)
    if unknown_fields:
        logger.warning(
            "ams_gtr %s: SODA schema drift -- %d unrecognised field(s) %s retained in "
            "immutable raw, dropped from the typed bronze projection.",
            dataset, len(unknown_fields), unknown_fields,
        )

    required = [spec.period_col] + sorted(spec.value_cols)
    missing = [c for c in required if c not in df.columns]
    if missing:
        # A value column absent from EVERY record is drift; Socrata omits null keys
        # per-record, so pandas only drops a column when no record carried it at all.
        raise ValueError(
            f"ams_gtr {dataset}: SODA payload is missing column(s) {missing}. "
            f"Present: {sorted(df.columns)}"
        )

    out = pd.DataFrame(index=df.index)
    out["period_date"] = df[spec.period_col].map(_to_date)

    if spec.key_col:
        out["route_or_reach"] = df[spec.key_col].astype("string")
    else:
        out["route_or_reach"] = pd.Series([pd.NA] * len(df), dtype="string")

    for source_col in spec.value_cols:
        unknown: set[str] = set()
        out[source_col] = [_to_float(v, unknown=unknown) for v in df[source_col]]
        out[source_col] = pd.to_numeric(out[source_col], errors="coerce").astype("float64")
        _warn_unknown_tokens(dataset, source_col, unknown)
        out[source_col] = _null_zero_sentinels(out[source_col], dataset, source_col)

    for extra in spec.extra_key_cols:
        if extra not in df.columns:
            raise ValueError(
                f"ams_gtr {dataset}: key column {extra!r} absent from the SODA payload."
            )
        out[extra] = df[extra].astype("string")

    null_periods = int(out["period_date"].isna().sum())
    if null_periods:
        logger.warning(
            "ams_gtr %s: dropping %d row(s) with an unparseable %s",
            dataset, null_periods, spec.period_col,
        )
        out = out.loc[out["period_date"].notna()].reset_index(drop=True)

    out["dataset"] = dataset
    out["as_of_date"] = as_of_date
    out["ingest_date"] = ingest_date
    out["source"] = SOURCE_NAME

    logger.info(
        "ams_gtr bronze (soda): dataset=%s rows=%d span=%s..%s as_of=%s",
        dataset, len(out),
        out["period_date"].min() if len(out) else "?",
        out["period_date"].max() if len(out) else "?",
        as_of_date,
    )
    return out


# ---------------------------------------------------------------------------
# GTRTable1.xlsx -> bronze
# ---------------------------------------------------------------------------

def transform_gtr_ocean_weekly_xlsx_to_bronze(
    raw_bytes: bytes,
    as_of_date: str,
    ingest_date: str,
) -> pd.DataFrame:
    """Parse ``GTRTable1.xlsx`` sheet ``Data`` into a typed bronze frame.

    The sheet is a working spreadsheet, not a data product: rows 1-3 are the
    analyst's own instructions, rows 4-5 are the current-week summary and its
    week-on-week deltas, row 6 is a group label, row 7 is the header, and the series
    starts at row 8.  Every one of those positions is ASSERTED before a value is
    read -- if AMS re-lays the sheet, this raises rather than silently reading the
    wrong column.

    Args:
        raw_bytes:   Raw bytes of the ``.xlsx`` file.
        as_of_date:  Snapshot date, ``YYYYMMDD``.
        ingest_date: ISO date (``YYYY-MM-DD``) this row was written.

    Returns:
        DataFrame with ``period_date`` plus Table 1's five value columns and the
        bronze metadata columns.  Source-faithful: the diesel / rail / river columns
        are kept even though silver publishes only the two ocean series.

    Raises:
        ValueError: If the sheet is absent or its header no longer matches.
        ImportError: If ``openpyxl`` is unavailable.
    """
    import io as _io

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment guard
        raise ImportError(
            "openpyxl is required to parse GTRTable1.xlsx (ams_gtr ocean_weekly)."
        ) from exc

    workbook = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=True, read_only=True)
    try:
        if _XLSX_SHEET not in workbook.sheetnames:
            raise ValueError(
                f"ams_gtr ocean_weekly: GTRTable1.xlsx has no {_XLSX_SHEET!r} sheet; "
                f"sheets present: {workbook.sheetnames}"
            )
        sheet = workbook[_XLSX_SHEET]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if len(rows) < _XLSX_FIRST_DATA_ROW:
        raise ValueError(
            f"ams_gtr ocean_weekly: sheet {_XLSX_SHEET!r} has only {len(rows)} rows; "
            f"the header alone sits on row {_XLSX_HEADER_ROW}."
        )

    def cell(row_1based: int, col_1based: int):
        row = rows[row_1based - 1]
        return row[col_1based - 1] if col_1based - 1 < len(row) else None

    # --- Assert the layout before reading a single number ---
    header_problems: list[str] = []
    for col, (expected, _bronze_name) in _XLSX_COLUMNS.items():
        observed = cell(_XLSX_HEADER_ROW, col)
        if str(observed).strip() != expected:
            header_problems.append(
                f"row {_XLSX_HEADER_ROW} col {col}: expected {expected!r}, got {observed!r}"
            )
    group_row, group_col, group_text = _XLSX_GROUP_CELL
    observed_group = cell(group_row, group_col)
    if str(observed_group).strip() != group_text:
        header_problems.append(
            f"row {group_row} col {group_col}: expected group label {group_text!r}, "
            f"got {observed_group!r}"
        )
    if header_problems:
        raise ValueError(
            "ams_gtr ocean_weekly: GTRTable1.xlsx sheet 'Data' no longer matches the "
            "pinned layout -- " + "; ".join(header_problems) + ". Re-read the sheet "
            "and re-pin _XLSX_COLUMNS before writing another row; reading the old "
            "positions against a re-laid sheet would publish the wrong series under "
            "the right name."
        )

    unknown_by_col: dict[str, set[str]] = {name: set() for _c, (_h, name) in _XLSX_COLUMNS.items()}
    records: list[dict] = []
    skipped_non_date = 0

    for row_index in range(_XLSX_FIRST_DATA_ROW, len(rows) + 1):
        period = _to_date_xlsx(cell(row_index, 1))
        if period is None:
            skipped_non_date += 1
            continue
        record: dict = {"period_date": period}
        for col, (_expected, bronze_name) in _XLSX_COLUMNS.items():
            if col == 1:
                continue
            record[bronze_name] = _to_float(
                cell(row_index, col), unknown=unknown_by_col[bronze_name]
            )
        records.append(record)

    if not records:
        raise ValueError(
            "ams_gtr ocean_weekly: GTRTable1.xlsx sheet 'Data' produced zero dated rows."
        )

    for name, unknown in unknown_by_col.items():
        _warn_unknown_tokens("ocean_weekly", name, unknown)

    if skipped_non_date:
        logger.info(
            "ams_gtr ocean_weekly: skipped %d non-dated row(s) below row %d "
            "(the sheet's own instruction and summary rows).",
            skipped_non_date, _XLSX_FIRST_DATA_ROW,
        )

    out = pd.DataFrame.from_records(records)
    for _col, (_expected, bronze_name) in _XLSX_COLUMNS.items():
        if bronze_name != "period_date":
            out[bronze_name] = pd.to_numeric(out[bronze_name], errors="coerce").astype("float64")
            out[bronze_name] = _null_zero_sentinels(out[bronze_name], "ocean_weekly", bronze_name)

    out["route_or_reach"] = pd.Series([pd.NA] * len(out), dtype="string")
    out["dataset"] = "ocean_weekly"
    out["as_of_date"] = as_of_date
    out["ingest_date"] = ingest_date
    out["source"] = SOURCE_NAME

    logger.info(
        "ams_gtr bronze (xlsx): dataset=ocean_weekly rows=%d span=%s..%s as_of=%s",
        len(out), out["period_date"].min(), out["period_date"].max(), as_of_date,
    )
    return out


def _to_date_xlsx(value) -> _dt.date | None:
    """Date coercion for spreadsheet cells, which may hold text or a datetime."""
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if len(text) < 10:
        return None
    try:
        return _dt.date.fromisoformat(text[:10])
    except ValueError:
        return None
