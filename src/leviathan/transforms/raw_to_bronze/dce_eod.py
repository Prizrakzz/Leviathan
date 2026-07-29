"""PRICE_AND_PLAYBOOKS W1c -- the DCE (Dalian) raw -> bronze transform. TWO parsers, ONE shape.

WHAT THIS MODULE OWNS
---------------------
The venue-specific half of ``silver_futures_eod``'s DCE leg, and nothing else:

  * :data:`DCE_VARIETY_MAP` -- the FIVE varieties this leg keeps and their leviathan slugs. It
    carries NO unit / currency / settle_kind / source: that authority is
    :mod:`leviathan.silver.futures_eod_contracts`, and an import-time assertion binds this map to
    the ``source == "dce"`` rows of ``CONTRACT_MAP`` in BOTH directions (the CZCE_ROOT_MAP
    precedent);
  * :func:`parse_dce_daily_json` -- the ``/dcereport/quote/delay/futureData?variety={v}`` body;
  * :func:`parse_dce_history_xlsx` -- the ``/dcereport/quote/history/download`` workbook;
  * the 4-digit ``YYMM`` contract-code decode, anchored on the SESSION the payload itself carries;
  * the ``0`` price sentinel, which both payload kinds use and which is the single most dangerous
    thing on this leg.

Both parsers emit the SAME :data:`BRONZE_COLUMNS` frame -- deliberately the CZCE column list,
column for column. Two Chinese exchanges, two wire formats, one bronze shape: the bronze -> silver
projection that lands later has one frame to read, and a daily row and a history row of the same
contract are directly comparable (which is how the backfill and the forward feed are reconciled at
their seam). ``root`` therefore keeps its CZCE name while carrying the DCE VARIETY letter.

Pure: pandas + openpyxl + the house logger. No boto3, no S3, no network, no playwright.

THE FOUR TRAPS THIS MODULE EXISTS TO NOT FALL INTO
--------------------------------------------------
1. **ZERO IS NOT A PRICE, IT IS THE UNDEFINED SENTINEL -- in BOTH payloads.** The daily API
   publishes ``settlePrice``/``closePrice`` of ``0.0`` for the whole board until the session closes
   (fixture ``dce_futureData_p.json`` was captured in exactly that state), and the history workbook
   prints ``"0"`` for open/high/low on a day a contract did not trade while STILL printing a real
   close and settle. A CNY/t contract cannot print 0, so every price cell equal to 0 is masked to
   NULL and counted. The mask is per-CELL and never touches ``volume`` / ``open_interest`` /
   ``oi_change`` / ``turnover``, whose zero is a true observation.
2. **An all-zero-settle daily payload is NOT_READY, not a session.** It is refused here
   (:func:`daily_not_ready` -> ValueError) as well as at the producer, which never lands it. The
   asymmetry matters: the night session rolls ``tradeDate`` FORWARD to T+1 while the settles are
   still zero, so an unguarded parse would write a whole zero-price board dated one day into the
   future -- plausible-looking rows for a session that has not happened.
3. **The history header is Chinese and is PINNED EXACTLY.** Unlike CZCE (whose header wording
   drifts across the years, forcing a positional-only map), the DCE workbook publishes a stable
   15-column header, so it is asserted verbatim and the parse fails CLOSED on any drift. The
   comparison is against :data:`HISTORY_HEADER` -- a variable of ``\\u`` escapes, so this SOURCE
   FILE stays pure ASCII and no code path ever prints the header (the Windows console is cp1252
   and a non-ASCII print crashes python). Drift is reported as escaped codepoints, never raw.
4. **The contract code is 4-digit ``YYMM`` and the century anchor is the PAYLOAD's own session.**
   ``p1601`` is January 2016 in the 2016 workbook and ``p2608`` is August 2026 in a 2026 capture;
   ``datetime.now()`` would silently re-date the whole backfill on any re-run. Same doctrine as the
   CZCE 3-digit decade anchor and the Databento single-digit year code.

SOURCE FIDELITY
---------------
No derived math. ``settle`` is the venue's own settlement field (``settlePrice`` / the 9th history
column), never the close; ``prev_settle`` is the venue's ``preSettlePrice`` / 8th column, never a
lag computed here. The daily API publishes no turnover and no open-interest change, so those two
columns are NULL on daily rows BY SOURCE -- not zero, which would be a fabricated count. There is
no FX conversion: a CNY/t settle stays CNY/t.
"""
from __future__ import annotations

import io
import json
import re
import warnings
from typing import Any, Optional

import pandas as pd

from leviathan.common.logging import get_logger
from leviathan.silver import futures_eod_contracts as FC

logger = get_logger(__name__)

# ---------------------------------------------------------------------------
# Varieties
# ---------------------------------------------------------------------------
# The five DCE varieties this leg keeps. The letter is the venue's own identity -- it is the
# ``variety=`` query parameter, the ``varietyId`` field, the raw key segment and the prefix of every
# contract code -- so it is what the map is keyed on. No unit / currency / settle_kind / source
# here on purpose; see the module docstring.
DCE_VARIETY_MAP: dict[str, str] = {
    "p": "palm_olein_dce",
    "a": "soybeans_no_1_dce",
    "b": "soybeans_no_2_dce",
    "m": "soybean_meal_dce",
    "y": "soybean_oil_dce",
}

# The publication ``source`` value these rows carry, verbatim from CONTRACT_MAP.
DCE_SOURCE = "dce"

# The oldest year the history download is walked from. Palm olein listed 2007-10 and the soy
# complex earlier; a year with no data answers with an empty/thin workbook rather than an error, so
# the walk is bounded rather than curated.
DCE_FIRST_HISTORY_YEAR = 2006

# A contract code: the variety letter(s) then FOUR digits of YYMM. Full anchors on both ends.
_CONTRACT_RE = re.compile(r"^([a-z]{1,2})([0-9]{4})$")

CONTRACT_MONTH_FMT = "%04d-%02d"

# The bronze frame's columns -- the CZCE list, verbatim. See the module docstring for why the
# variety letter lives in a column called ``root``.
BRONZE_COLUMNS: list[str] = [
    "trade_date", "leviathan_slug", "root", "raw_symbol", "contract_month",
    "prev_settle", "open", "high", "low", "close", "settle",
    "volume", "open_interest", "oi_change", "turnover",
]

# Zero is not a possible CNY/t price. These are the columns the sentinel mask applies to, and the
# ones it must NEVER apply to are the counts (volume, open_interest, oi_change, turnover).
_PRICE_COLUMNS = ("prev_settle", "open", "high", "low", "close", "settle")
_UNDEFINED_PRICE = 0.0


def _lint_variety_map() -> list[str]:
    """DCE_VARIETY_MAP must be EXACTLY the ``source == 'dce'`` slugs of CONTRACT_MAP, both ways."""
    errs: list[str] = []
    mapped = set(DCE_VARIETY_MAP.values())
    curated = {slug for slug, rec in FC.CONTRACT_MAP.items() if rec["source"] == DCE_SOURCE}
    for slug in sorted(mapped - curated):
        errs.append(f"{slug}: in DCE_VARIETY_MAP but not a source={DCE_SOURCE!r} CONTRACT_MAP slug")
    for slug in sorted(curated - mapped):
        errs.append(f"{slug}: a source={DCE_SOURCE!r} CONTRACT_MAP slug with no DCE variety")
    for variety in sorted(DCE_VARIETY_MAP):
        if not _CONTRACT_RE.match(variety + "0000"):
            errs.append(f"{variety!r}: not a 1-2 character lowercase DCE variety letter")
    if len(set(DCE_VARIETY_MAP.values())) != len(DCE_VARIETY_MAP):
        errs.append("two varieties map to the same leviathan slug")
    return errs


# Import-time fail-closed, the CZCE_ROOT_MAP / yfinance UNIT_MAP precedent: a variety map that has
# drifted from the curated contract map must never reach a producer.
assert not _lint_variety_map(), \
    "dce_eod.DCE_VARIETY_MAP is malformed: " + "; ".join(_lint_variety_map())


# ---------------------------------------------------------------------------
# Shared decode
# ---------------------------------------------------------------------------
def ascii_safe(value: Any, limit: int = 300) -> str:
    """``value`` as pure ASCII escapes. EVERY error message that can carry venue text uses this.

    The DCE payloads are full of Chinese (the ``msg`` field, the commodity-name column, the sheet
    name), the Windows console is cp1252, and a non-ASCII log record crashes python there. Escaping
    at the boundary is what lets a header-drift error still SHOW the drift."""
    text = str(value)
    if len(text) > limit:
        text = text[:limit] + "..."
    return text.encode("ascii", "backslashreplace").decode("ascii")


def parse_number(token: Any) -> float:
    """``'4,580' -> 4580.0``; blank / '-' / unparseable -> NaN. Comma thousands, no locale.

    The history workbook stores EVERY cell as an inline string with comma separators (the daily API
    sends real JSON numbers), so both payload kinds funnel through here."""
    if token is None or isinstance(token, bool):
        return float("nan")
    if isinstance(token, (int, float)):
        return float(token)
    tok = str(token).strip().replace(",", "").replace(" ", "")
    if not tok or tok in {"-", "--"}:
        return float("nan")
    try:
        return float(tok)
    except ValueError:
        return float("nan")


def resolve_contract_year(year_2digit: int, trade_date: str) -> int:
    """The SESSION-ANCHORED century rule for the 2-digit year of a ``YYMM`` contract code.

    Of the three candidate years whose final two digits match and which sit within a century of the
    session, take the nearest to the session year, preferring the FUTURE on a tie. DCE lists at most
    ~14 months of delivery months, so the nearest candidate is the only reachable one, and the rule
    stays correct across a century boundary in both directions.

    The anchor is the session date the PAYLOAD carries, never ``datetime.now()``: a re-run in 2031
    over the 2016 raw bytes must decode identically. Same doctrine as the CZCE decade anchor."""
    yy = int(year_2digit)
    if not 0 <= yy <= 99:
        raise ValueError(f"year_2digit {year_2digit!r} is not a two-digit year")
    ts = pd.Timestamp(trade_date)
    base = ts.year - (ts.year % 100) + yy
    best, best_key = None, None
    for cand in (base - 100, base, base + 100):
        key = (abs(cand - ts.year), 0 if cand >= ts.year else 1)   # tie -> prefer the future
        if best_key is None or key < best_key:
            best, best_key = cand, key
    return int(best)


def contract_month_str(code_digits: str, trade_date: str) -> str:
    """``('2608', '2026-07-30') -> '2026-08'``. Fail-closed on a month outside 1..12."""
    digits = str(code_digits)
    if len(digits) != 4 or not digits.isdigit():
        raise ValueError(f"{code_digits!r} is not a 4-digit DCE YYMM contract code")
    month = int(digits[2:])
    if not 1 <= month <= 12:
        raise ValueError(f"{code_digits!r}: delivery month {month} is outside 1..12")
    return CONTRACT_MONTH_FMT % (resolve_contract_year(int(digits[:2]), trade_date), month)


def split_contract_id(contract_id: str, *, variety: Optional[str] = None) -> tuple[str, str]:
    """``'p2608' -> ('p', '2608')``. Fail-closed on shape, on an unmapped variety, and -- when
    ``variety`` is supplied -- on a code that belongs to a DIFFERENT variety than the object claims.

    That last check is the one that matters: the daily endpoint is per-variety and the raw key
    carries the variety, so a body served for the wrong variety would otherwise land palm-olein
    prices under the soybean-oil slug, which is a plausible WRONG number rather than an error."""
    code = str(contract_id).strip()
    m = _CONTRACT_RE.match(code)
    if not m:
        raise ValueError(f"{ascii_safe(code)!r} is not a DCE contract code (variety letter + YYMM)")
    got, digits = m.group(1), m.group(2)
    if got not in DCE_VARIETY_MAP:
        raise ValueError(
            f"contract {code!r} carries variety {got!r}, which is not one of the five this leg "
            f"keeps {sorted(DCE_VARIETY_MAP)} -- refusing to guess a slug")
    if variety is not None and got != variety:
        raise ValueError(
            f"contract {code!r} is variety {got!r} but the object is variety {variety!r} -- the "
            f"venue served a different board; refusing to parse (this is how a palm-olein price "
            f"lands under the soybean-oil slug)")
    return got, digits


def _mask_price_sentinels(values: dict) -> int:
    """Mask every 0 PRICE cell to NULL, in place; return how many were masked."""
    masked = 0
    for col in _PRICE_COLUMNS:
        if values.get(col) == _UNDEFINED_PRICE:
            values[col] = float("nan")
            masked += 1
    return masked


def _finalize(rows: list[dict]) -> pd.DataFrame:
    """The shared bronze frame build -- identical dtypes from both parsers."""
    df = pd.DataFrame(rows, columns=BRONZE_COLUMNS)
    if len(df):
        df["trade_date"] = pd.to_datetime(df["trade_date"]).astype("datetime64[us]")
        for col in ("volume", "open_interest", "oi_change"):
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    return df


# ---------------------------------------------------------------------------
# The DAILY quote API
# ---------------------------------------------------------------------------
_ENVELOPE_CODE_OK = 200


def daily_records(payload: bytes) -> tuple[list[dict], dict]:
    """``(contract records, envelope)`` from one ``futureData`` body. Fail-closed on the envelope.

    The endpoint answers HTTP 200 with ``{"success": false, "code": ...}`` for a bad variety, so
    the envelope is checked rather than the transport status."""
    text = payload.decode("utf-8") if isinstance(payload, (bytes, bytearray)) else str(payload)
    obj = json.loads(text)
    if not isinstance(obj, dict):
        raise ValueError(f"dce daily: body is {type(obj).__name__}, expected a JSON object")
    envelope = {k: v for k, v in obj.items() if k != "data"}
    if obj.get("success") is False or int(obj.get("code") or 0) != _ENVELOPE_CODE_OK:
        raise ValueError(
            f"dce daily: the API refused the request (success={obj.get('success')!r}, "
            f"code={obj.get('code')!r}, msg={ascii_safe(obj.get('msg'))!r})")
    data = obj.get("data")
    if not isinstance(data, list):
        raise ValueError(f"dce daily: 'data' is {type(data).__name__}, expected a list")
    return [r for r in data if isinstance(r, dict)], envelope


def daily_not_ready(records: list[dict]) -> bool:
    """True when the board is NOT SETTLED YET -- every contract's ``settlePrice`` is the 0 sentinel.

    This is the guard the producer fires on (skip the variety, never land the object) and the guard
    the parser refuses on. An empty board counts as not ready: nothing to settle is not a session.

    It is deliberately ALL-contracts and not any: a single thin contract legitimately settles at a
    real price while never trading, and a board where one contract reads 0 while the rest are real
    is a per-contract sentinel (masked to NULL by the row parse), not a timing problem."""
    if not records:
        return True
    return all(parse_number(r.get("settlePrice")) == _UNDEFINED_PRICE for r in records)


def parse_dce_daily_json(payload: bytes, *, variety: Optional[str] = None,
                         as_of_date: Optional[str] = None) -> tuple[pd.DataFrame, dict]:
    """One ``futureData?variety={v}`` body -> bronze rows + a stats dict.

    ``variety`` is the raw key's own ``variety=`` segment when the caller has it; it is CHECKED
    against every contract code rather than trusted (see :func:`split_contract_id`). ``as_of_date``
    is the CAPTURE date and is carried into the stats for audit -- it is NEVER the trade date: the
    payload's own ``tradeDate`` is the session, exactly as the JSE sheet's header date is."""
    records, envelope = daily_records(payload)
    if daily_not_ready(records):
        raise ValueError(
            f"dce daily variety={variety!r}: every one of {len(records)} contract(s) carries "
            f"settlePrice 0.0 -- the board has NOT settled yet (the night session rolls tradeDate "
            f"to T+1 with zero settles). This object should never have been landed; refusing to "
            f"write a zero-price board")

    sessions: set[str] = set()
    rows: list[dict] = []
    zero_price_cells = 0
    for rec in records:
        code = str(rec.get("contractId") or "").strip()
        got_variety, digits = split_contract_id(code, variety=variety)
        session = _iso_from_compact(rec.get("tradeDate"), context=f"dce daily {code}")
        sessions.add(session)
        values = {
            "prev_settle": parse_number(rec.get("preSettlePrice")),
            "open": parse_number(rec.get("openPrice")),
            "high": parse_number(rec.get("highPrice")),
            "low": parse_number(rec.get("lowPrice")),
            "close": parse_number(rec.get("closePrice")),
            "settle": parse_number(rec.get("settlePrice")),
            "volume": parse_number(rec.get("volume")),
            "open_interest": parse_number(rec.get("openInterest")),
            # NOT published by this endpoint. NULL by source -- never 0, which would be a count
            # this leg invented.
            "oi_change": float("nan"),
            "turnover": float("nan"),
        }
        zero_price_cells += _mask_price_sentinels(values)
        rows.append({
            "trade_date": session,
            "leviathan_slug": DCE_VARIETY_MAP[got_variety],
            "root": got_variety,
            "raw_symbol": code,               # VERBATIM. Never parsed into meaning at ingest.
            "contract_month": contract_month_str(digits, session),
            **values,
        })

    if len(sessions) > 1:
        raise ValueError(
            f"dce daily variety={variety!r}: the body carries {len(sessions)} distinct tradeDate "
            f"values {sorted(sessions)} -- one capture is one board; refusing to parse")
    df = _finalize(rows)
    stats = {
        "kind": "daily",
        "variety": variety,
        "as_of_date": as_of_date,
        "trade_date": sorted(sessions)[0] if sessions else None,
        "contracts": len(records),
        "rows_kept": int(len(df)),
        "zero_price_cells": zero_price_cells,
        "request_id": envelope.get("requestId"),
    }
    logger.info("dce daily bronze variety=%s %s: %d row(s), %d zero-price cell(s) masked",
                variety, stats["trade_date"], len(df), zero_price_cells)
    return df, stats


def _iso_from_compact(value: Any, *, context: str) -> str:
    """``'20160104' -> '2016-01-04'``. Fail-closed; there is no wall-clock fallback anywhere here."""
    token = str(value or "").strip().replace("-", "")
    if len(token) != 8 or not token.isdigit():
        raise ValueError(f"{context}: {ascii_safe(value)!r} is not a YYYYMMDD trade date")
    return f"{token[:4]}-{token[4:6]}-{token[6:]}"


# ---------------------------------------------------------------------------
# The HISTORY workbook
# ---------------------------------------------------------------------------
# THE PINNED HEADER, column A..O, as \u escapes so this file stays pure ASCII (cp1252 console).
# Meanings, in order:
#   0 commodity name   1 contract       2 trade date     3 open          4 high
#   5 low              6 close          7 prev settle    8 SETTLE        9 change
#  10 change1         11 volume        12 open interest 13 OI change    14 turnover
HISTORY_HEADER: tuple[str, ...] = (
    "\u5546\u54c1\u540d\u79f0",               # commodity name
    "\u5408\u7ea6\u540d\u79f0",               # contract
    "\u4ea4\u6613\u65e5\u671f",               # trade date (YYYYMMDD)
    "\u5f00\u76d8\u4ef7",                     # open
    "\u6700\u9ad8\u4ef7",                     # high
    "\u6700\u4f4e\u4ef7",                     # low
    "\u6536\u76d8\u4ef7",                     # close
    "\u524d\u7ed3\u7b97\u4ef7",               # prev settle
    "\u7ed3\u7b97\u4ef7",                     # SETTLE
    "\u6da8\u8dcc",                           # change
    "\u6da8\u8dcc1",                          # change1
    "\u6210\u4ea4\u91cf",                     # volume
    "\u6301\u4ed3\u91cf",                     # open interest
    "\u6301\u4ed3\u91cf\u53d8\u5316",         # open-interest change
    "\u6210\u4ea4\u989d",                     # turnover
)
# Positional map, resolved ONLY after the header assertion above passes.
_H_CONTRACT = 1
_H_TRADE_DATE = 2
_H_OPEN = 3
_H_HIGH = 4
_H_LOW = 5
_H_CLOSE = 6
_H_PREV_SETTLE = 7
_H_SETTLE = 8
_H_VOLUME = 11
_H_OPEN_INTEREST = 12
_H_OI_CHANGE = 13
_H_TURNOVER = 14
HISTORY_COLUMN_COUNT = len(HISTORY_HEADER)


def read_history_grid(payload: bytes) -> tuple[list[list], str]:
    """``(rows, sheet_name)`` from one DCE history workbook.

    ``openpyxl`` in read-only mode. Every cell in this workbook is an ``inlineStr`` (the
    sharedStrings table is empty), so every value arrives as a STRING with comma thousands
    separators -- which is why nothing downstream trusts a cell's python type.

    Split from the parse at the workbook boundary on purpose (the JSE ``read_grid`` precedent):
    everything interesting is grid logic, and the tests exercise it against the real vendor bytes."""
    import openpyxl  # lazy: keeps the module importable in a stripped environment

    with warnings.catch_warnings():
        # The vendor workbook is machine-generated and carries no default style; openpyxl warns and
        # then reads it fine. The warning is noise on every single history unit.
        warnings.simplefilter("ignore", UserWarning)
        book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        try:
            sheet = book[book.sheetnames[0]]
            rows = [list(r) for r in sheet.iter_rows(values_only=True)]
            name = str(book.sheetnames[0])
        finally:
            book.close()
    return rows, name


def assert_history_header(row: list) -> None:
    """The fail-closed header pin. Raises with ESCAPED codepoints on any drift.

    Unlike CZCE -- whose header wording drifts across the history, which is why that parse is
    positional-only -- the DCE workbook publishes a stable header, so it is asserted verbatim and
    the positional map below is only used once this passes. The error message escapes both sides so
    it is printable on a cp1252 console; the header itself is NEVER printed raw anywhere."""
    got = tuple("" if c is None else str(c).strip() for c in (row or []))
    if got == HISTORY_HEADER:
        return
    diffs = []
    for i in range(max(len(got), HISTORY_COLUMN_COUNT)):
        mine = HISTORY_HEADER[i] if i < HISTORY_COLUMN_COUNT else "<absent>"
        theirs = got[i] if i < len(got) else "<absent>"
        if mine != theirs:
            diffs.append(f"col{i}: got {ascii_safe(theirs, 60)!r} want {ascii_safe(mine, 60)!r}")
    raise ValueError(
        f"dce history: the header row drifted ({len(got)} column(s), expected "
        f"{HISTORY_COLUMN_COUNT}). {'; '.join(diffs[:8])}. The positional map cannot be trusted "
        f"through a header change -- refusing to parse (fail closed, never a positional guess)")


def parse_dce_history_xlsx(payload: bytes, *, variety: Optional[str] = None,
                           year: Optional[int] = None) -> tuple[pd.DataFrame, dict]:
    """One ``{v}_ftr.xlsx`` year workbook -> bronze rows + a stats dict.

    Each row carries its OWN trade date (the workbook is one variety x one calendar year, every
    listed contract x every session), so the contract-code century anchor is per row -- there is no
    single session date on this payload and none is invented."""
    grid, sheet_name = read_history_grid(payload)
    if not grid:
        raise ValueError("dce history: the workbook is empty")
    assert_history_header(grid[0])

    rows: list[dict] = []
    contracts: set[str] = set()
    sessions: set[str] = set()
    zero_price_cells = 0
    varieties: set[str] = set()
    for idx, raw in enumerate(grid[1:], start=2):
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue                       # a trailing blank row; not a data row
        if len(raw) < HISTORY_COLUMN_COUNT:
            raise ValueError(
                f"dce history row {idx}: {len(raw)} cell(s), expected {HISTORY_COLUMN_COUNT} -- "
                f"the positional map cannot be trusted on this row")
        code = str(raw[_H_CONTRACT] or "").strip()
        got_variety, digits = split_contract_id(code, variety=variety)
        varieties.add(got_variety)
        session = _iso_from_compact(raw[_H_TRADE_DATE], context=f"dce history row {idx}")
        contracts.add(code)
        sessions.add(session)
        values = {
            "prev_settle": parse_number(raw[_H_PREV_SETTLE]),
            "open": parse_number(raw[_H_OPEN]),
            "high": parse_number(raw[_H_HIGH]),
            "low": parse_number(raw[_H_LOW]),
            "close": parse_number(raw[_H_CLOSE]),
            "settle": parse_number(raw[_H_SETTLE]),
            "volume": parse_number(raw[_H_VOLUME]),
            "open_interest": parse_number(raw[_H_OPEN_INTEREST]),
            "oi_change": parse_number(raw[_H_OI_CHANGE]),
            "turnover": parse_number(raw[_H_TURNOVER]),
        }
        # The untraded-day sentinel: open/high/low (and volume) print "0" while close and settle
        # print real values. Prices go NULL; the counts stay 0, because a zero volume IS the
        # observation that the contract did not trade.
        zero_price_cells += _mask_price_sentinels(values)
        rows.append({
            "trade_date": session,
            "leviathan_slug": DCE_VARIETY_MAP[got_variety],
            "root": got_variety,
            "raw_symbol": code,               # VERBATIM. Never parsed into meaning at ingest.
            "contract_month": contract_month_str(digits, session),
            **values,
        })

    df = _finalize(rows)
    stats = {
        "kind": "history",
        "variety": variety or (sorted(varieties)[0] if varieties else None),
        "year": int(year) if year is not None else None,
        "sheet": ascii_safe(sheet_name, 80),
        "grid_rows": len(grid),
        "rows_kept": int(len(df)),
        "contracts": len(contracts),
        "sessions": len(sessions),
        "first_trade_date": min(sessions) if sessions else None,
        "last_trade_date": max(sessions) if sessions else None,
        "zero_price_cells": zero_price_cells,
    }
    logger.info("dce history bronze variety=%s year=%s: %d row(s), %d contract(s), %d session(s), "
                "%d zero-price cell(s) masked", stats["variety"], stats["year"], len(df),
                len(contracts), len(sessions), zero_price_cells)
    return df, stats


# ---------------------------------------------------------------------------
# The one entry point the silver task calls
# ---------------------------------------------------------------------------
def build_dce_bronze(payload: bytes, *, variety: Optional[str] = None,
                     kind: str = "daily", as_of_date: Optional[str] = None,
                     year: Optional[int] = None) -> tuple[pd.DataFrame, dict]:
    """Dispatch one landed DCE object to its parser. ``kind`` is decided by the KEY, not sniffed.

    (``history/`` is a path segment, so the reader always knows which payload it holds; sniffing the
    bytes would make a truncated download look like the other format.)"""
    if kind == "history":
        return parse_dce_history_xlsx(payload, variety=variety, year=year)
    if kind == "daily":
        return parse_dce_daily_json(payload, variety=variety, as_of_date=as_of_date)
    raise ValueError(f"dce: unknown payload kind {kind!r} (expected 'daily' or 'history')")
