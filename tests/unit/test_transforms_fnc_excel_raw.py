"""Unit tests for FNC Colombia Excel raw -> bronze parsing."""
from __future__ import annotations

import io
from datetime import date

from leviathan.transforms.raw_to_bronze.fnc_excel import extract_fnc_excel
from openpyxl import Workbook


def _add_monthly_sheet(wb: Workbook, name: str, header: str, value: float) -> None:
    ws = wb.create_sheet(name)
    ws["C2"] = header
    ws["D6"] = "Mes"
    ws["E6"] = "Valor"
    ws["D7"] = date(2024, 1, 1)
    ws["E7"] = value
    ws["D8"] = date(2024, 2, 1)
    ws["E8"] = value + 1


def _workbook_bytes() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _add_monthly_sheet(wb, "8. Producción mensual", "Produccion", 700.0)
    _add_monthly_sheet(wb, "3. Precio Ex_Dock Mensual", "Ex dock", 310.5)
    _add_monthly_sheet(wb, "2. Precio Interno Mensual", "Interno", 1_950_000.0)
    _add_monthly_sheet(wb, "1. Total_Volumen", "Volumen", 925.0)
    _add_monthly_sheet(wb, "2. Total_Valor", "Valor", 363.3)

    area = wb.create_sheet("7. Área cult. dep. producto")
    area["B6"] = "Departamento"
    area["C6"] = 2024
    area["D6"] = "2025*"
    area["B7"] = "Nariño"
    area["C7"] = 35.2
    area["D7"] = 36.1

    port = wb.create_sheet("5. Puerto_Tipo_Vol_Val")
    headers = [
        "Año",
        "Mes",
        "Puerto de embarque",
        "Tipo de café",
        "Sacos de 70 kg. equivalente real Exportados",
        "Sacos de 60 Kg. Exportados",
        "Valor provisional de la Exportación (USD) *",
    ]
    for col, header in enumerate(headers, start=3):
        port.cell(row=8, column=col, value=header)
    values = [2024, 1, "Aerp. El Dorado", "Café Verde", 74, 84, 60776.73]
    for col, value in enumerate(values, start=3):
        port.cell(row=9, column=col, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extracts_all_expected_fnc_series() -> None:
    series = extract_fnc_excel(_workbook_bytes(), "fnc.xlsx", "2026-06-02")
    assert set(series) == {
        "produccion_mensual",
        "precio_ex_dock_mensual",
        "precio_interno_mensual",
        "area_departamento",
        "exportaciones_total_volumen",
        "exportaciones_total_valor",
        "exportaciones_puerto_tipo",
    }


def test_parses_monthly_date_value_sheets() -> None:
    series = extract_fnc_excel(_workbook_bytes(), "fnc.xlsx", "2026-06-02")
    production = series["produccion_mensual"]
    assert list(production.columns) == [
        "series_name",
        "year",
        "month",
        "date",
        "value",
        "unit",
        "source_file",
        "source",
        "ingest_date",
    ]
    assert len(production) == 2
    assert production.iloc[0]["year"] == 2024
    assert production.iloc[0]["month"] == 1
    assert production.iloc[0]["date"] == date(2024, 1, 1)
    assert production.iloc[0]["unit"] == "1000_bags_60kg"


def test_parses_area_department_and_normalizes_accents() -> None:
    series = extract_fnc_excel(_workbook_bytes(), "fnc.xlsx", "2026-06-02")
    area = series["area_departamento"]
    assert len(area) == 2
    assert set(area["department"]) == {"narino"}
    assert set(area["year"]) == {2024, 2025}
    assert area.iloc[0]["area_1000_ha"] == 35.2


def test_parses_export_port_type_sheet() -> None:
    series = extract_fnc_excel(_workbook_bytes(), "fnc.xlsx", "2026-06-02")
    exports = series["exportaciones_puerto_tipo"]
    row = exports.iloc[0]
    assert row["year"] == 2024
    assert row["month"] == 1
    assert row["port"] == "aerp_el_dorado"
    assert row["coffee_type"] == "cafe_verde"
    assert row["exports_bags_60kg"] == 84.0
    assert row["exports_value_usd"] == 60776.73
